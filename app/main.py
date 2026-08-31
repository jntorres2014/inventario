from __future__ import annotations

import json
import os
import re
import sqlite3
import unicodedata
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import PatternFill
from pydantic import BaseModel, Field

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = Path(os.getenv("INVENTARIO_DATA_DIR", BASE_DIR / "data"))
DB_PATH = DATA_DIR / "inventario.db"
SOURCE_PATH = DATA_DIR / "inventario_origen.xlsx"
EXPORT_PATH = DATA_DIR / "inventario_resultado.xlsx"
STATIC_DIR = BASE_DIR / "app" / "static"

GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")

class ScanRequest(BaseModel):
    code: str = Field(min_length=1, max_length=200)
    scanner_name: str = Field(default="Sin identificar", max_length=100)


def now_iso() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")


def connect() -> sqlite3.Connection:
    connection = sqlite3.connect(DB_PATH, timeout=10)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA journal_mode=WAL")
    return connection


def init_db() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with connect() as connection:
        connection.executescript(
            """
            CREATE TABLE IF NOT EXISTS items (
                code TEXT PRIMARY KEY,
                row_number INTEGER NOT NULL,
                found INTEGER NOT NULL DEFAULT 0,
                first_scanned_at TEXT,
                scanner_name TEXT,
                scan_count INTEGER NOT NULL DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS unknown_scans (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL,
                scanned_at TEXT NOT NULL,
                scanner_name TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS metadata (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );
            """
        )


@asynccontextmanager
async def lifespan(_: FastAPI):
    init_db()
    yield


app = FastAPI(title="Inventario local", version="0.1.0", lifespan=lifespan)
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


def normalize_header(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]", "", text.lower())


def normalize_code(value: Any) -> str:
    text = str(value or "").strip()
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return text


def cell_code(cell: Cell) -> str:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        number_format = str(cell.number_format or "")
        zero_format = re.fullmatch(r"0+", number_format)
        if zero_format and isinstance(value, int):
            return f"{value:0{len(number_format)}d}"
    return normalize_code(value)


def find_inventory_column(headers: list[Any], requested: str | None) -> int:
    if requested:
        requested_normalized = normalize_header(requested)
        for index, header in enumerate(headers, start=1):
            if normalize_header(header) == requested_normalized:
                return index
        raise ValueError(f"No existe la columna '{requested}'.")

    candidates = {
        "inventario",
        "numeroinventario",
        "nroinventario",
        "nrodeinventario",
        "nodeinventario",
        "codigoinventario",
        "codigodeinventario",
    }
    for index, header in enumerate(headers, start=1):
        if normalize_header(header) in candidates:
            return index
    raise ValueError("No se pudo detectar automáticamente la columna de inventario.")


def store_metadata(connection: sqlite3.Connection, key: str, value: Any) -> None:
    connection.execute(
        "INSERT INTO metadata(key, value) VALUES(?, ?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (key, json.dumps(value, ensure_ascii=False)),
    )


def read_metadata(connection: sqlite3.Connection, key: str, default: Any = None) -> Any:
    row = connection.execute("SELECT value FROM metadata WHERE key = ?", (key,)).fetchone()
    return json.loads(row["value"]) if row else default


@app.get("/")
def index() -> FileResponse:
    return FileResponse(STATIC_DIR / "index.html")


@app.post("/api/inventory/import")
async def import_inventory(
    file: UploadFile = File(...), inventory_column: str = Form(default="")
) -> dict[str, Any]:
    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Por ahora se requiere un archivo .xlsx.")

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    content = await file.read()
    if len(content) > 25 * 1024 * 1024:
        raise HTTPException(400, "El archivo supera el máximo de 25 MB.")
    SOURCE_PATH.write_bytes(content)

    try:
        workbook = load_workbook(SOURCE_PATH, data_only=False)
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]
        column_index = find_inventory_column(headers, inventory_column or None)
    except ValueError as exc:
        SOURCE_PATH.unlink(missing_ok=True)
        raise HTTPException(
            400,
            {"message": str(exc), "available_columns": [str(h) for h in headers if h]},
        ) from exc
    except Exception as exc:
        SOURCE_PATH.unlink(missing_ok=True)
        raise HTTPException(400, "No se pudo leer el archivo Excel.") from exc

    items: list[tuple[str, int]] = []
    duplicates: list[str] = []
    seen: set[str] = set()
    for row_number in range(2, worksheet.max_row + 1):
        code = cell_code(worksheet.cell(row=row_number, column=column_index))
        if not code:
            continue
        if code in seen:
            duplicates.append(code)
            continue
        seen.add(code)
        items.append((code, row_number))

    if not items:
        SOURCE_PATH.unlink(missing_ok=True)
        raise HTTPException(400, "La columna seleccionada no contiene números de inventario.")

    with connect() as connection:
        connection.execute("BEGIN IMMEDIATE")
        connection.execute("DELETE FROM items")
        connection.execute("DELETE FROM unknown_scans")
        connection.execute("DELETE FROM metadata")
        connection.executemany("INSERT INTO items(code, row_number) VALUES(?, ?)", items)
        store_metadata(connection, "filename", filename)
        store_metadata(connection, "sheet_name", worksheet.title)
        store_metadata(connection, "column_index", column_index)
        store_metadata(connection, "column_name", headers[column_index - 1])
        store_metadata(connection, "duplicate_codes", duplicates)

    return {
        "ok": True,
        "total": len(items),
        "column": headers[column_index - 1],
        "duplicates_in_file": duplicates,
    }


@app.post("/api/scan")
def scan(request: ScanRequest) -> dict[str, Any]:
    code = normalize_code(request.code)
    scanner_name = request.scanner_name.strip() or "Sin identificar"
    scanned_at = now_iso()
    if not code:
        raise HTTPException(400, "El código está vacío.")

    with connect() as connection:
        connection.execute("BEGIN IMMEDIATE")
        item = connection.execute("SELECT * FROM items WHERE code = ?", (code,)).fetchone()
        if item is None:
            connection.execute(
                "INSERT INTO unknown_scans(code, scanned_at, scanner_name) VALUES(?, ?, ?)",
                (code, scanned_at, scanner_name),
            )
            return {"status": "unknown", "code": code, "message": "Código no registrado"}

        repeated = bool(item["found"])
        connection.execute(
            """
            UPDATE items
            SET found = 1,
                first_scanned_at = COALESCE(first_scanned_at, ?),
                scanner_name = COALESCE(scanner_name, ?),
                scan_count = scan_count + 1
            WHERE code = ?
            """,
            (scanned_at, scanner_name, code),
        )
        return {
            "status": "repeated" if repeated else "found",
            "code": code,
            "message": "Código repetido" if repeated else "Equipo encontrado",
            "scan_count": int(item["scan_count"]) + 1,
        }


@app.get("/api/stats")
def stats() -> dict[str, Any]:
    with connect() as connection:
        total = connection.execute("SELECT COUNT(*) FROM items").fetchone()[0]
        found = connection.execute("SELECT COUNT(*) FROM items WHERE found = 1").fetchone()[0]
        unknown = connection.execute("SELECT COUNT(*) FROM unknown_scans").fetchone()[0]
        duplicates = connection.execute(
            "SELECT COALESCE(SUM(scan_count - 1), 0) FROM items WHERE scan_count > 1"
        ).fetchone()[0]
        filename = read_metadata(connection, "filename")
    return {
        "loaded": total > 0,
        "filename": filename,
        "total": total,
        "found": found,
        "pending": total - found,
        "unknown": unknown,
        "duplicates": duplicates,
        "percent": round((found / total * 100), 1) if total else 0,
    }


@app.get("/api/items")
def list_items(status: str = "all", limit: int = 100) -> dict[str, Any]:
    limit = max(1, min(limit, 500))
    where = ""
    if status == "found":
        where = "WHERE found = 1"
    elif status == "pending":
        where = "WHERE found = 0"
    with connect() as connection:
        rows = connection.execute(
            f"SELECT code, found, first_scanned_at, scanner_name, scan_count "
            f"FROM items {where} ORDER BY row_number LIMIT ?",
            (limit,),
        ).fetchall()
    return {"items": [dict(row) for row in rows]}


@app.get("/api/export")
def export_inventory() -> FileResponse:
    if not SOURCE_PATH.exists():
        raise HTTPException(404, "Todavía no se cargó un inventario.")

    workbook = load_workbook(SOURCE_PATH)
    worksheet = workbook.active
    with connect() as connection:
        items = connection.execute("SELECT * FROM items").fetchall()
        unknown = connection.execute("SELECT * FROM unknown_scans ORDER BY id").fetchall()
        column_index = read_metadata(connection, "column_index")

    first_new_column = worksheet.max_column + 1
    headers = ["Inventariado", "Fecha de inventario", "Escaneado por", "Cantidad de lecturas"]
    for offset, header in enumerate(headers):
        worksheet.cell(1, first_new_column + offset, header)

    for item in items:
        row = item["row_number"]
        found = bool(item["found"])
        values = [
            "SÍ" if found else "NO",
            item["first_scanned_at"] or "",
            item["scanner_name"] or "",
            item["scan_count"],
        ]
        for offset, value in enumerate(values):
            worksheet.cell(row, first_new_column + offset, value)
        worksheet.cell(row, column_index).fill = GREEN_FILL if found else RED_FILL

    if "Códigos no registrados" in workbook.sheetnames:
        del workbook["Códigos no registrados"]
    unknown_sheet = workbook.create_sheet("Códigos no registrados")
    unknown_sheet.append(["Código", "Fecha", "Escaneado por"])
    for row in unknown:
        unknown_sheet.append([row["code"], row["scanned_at"], row["scanner_name"]])

    workbook.save(EXPORT_PATH)
    return FileResponse(
        EXPORT_PATH,
        filename="inventario_resultado.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/api/demo")
def create_demo() -> dict[str, Any]:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Activos"
    worksheet.append(["Nro Inventario", "Descripción", "Área"])
    worksheet.append(["000101", "Notebook de prueba", "Sistemas"])
    worksheet.append(["000102", "Monitor de prueba", "Sistemas"])
    worksheet.append(["000103", "Impresora de prueba", "Administración"])
    workbook.save(SOURCE_PATH)

    with connect() as connection:
        connection.execute("BEGIN IMMEDIATE")
        connection.execute("DELETE FROM items")
        connection.execute("DELETE FROM unknown_scans")
        connection.execute("DELETE FROM metadata")
        connection.executemany(
            "INSERT INTO items(code, row_number) VALUES(?, ?)",
            [("000101", 2), ("000102", 3), ("000103", 4)],
        )
        store_metadata(connection, "filename", "inventario_demo.xlsx")
        store_metadata(connection, "sheet_name", "Activos")
        store_metadata(connection, "column_index", 1)
        store_metadata(connection, "column_name", "Nro Inventario")
    return {"ok": True, "total": 3, "codes": ["000101", "000102", "000103"]}
